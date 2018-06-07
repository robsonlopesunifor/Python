import unittest
import pandas as pd
import numpy as np
import openpyxl

class OpenPyXl(unittest.TestCase):

    def test_importar(self):
        wb = openpyxl.load_workbook('G:\Meu Drive\DA  Software\Python\OpenPyXl - Execel\Execel\example.xlsx')
        print type(wb)

    def test_obtendo_folhas_da_pasta(self):
        print ''
        print '-------------obtendo folhas da pasta-------------------'
        wb = openpyxl.load_workbook('G:\Meu Drive\DA  Software\Python\OpenPyXl - Execel\Execel\example.xlsx')
        print type(wb)
        print "------nome das folhas        ['Sheet1', 'Sheet2', 'Sheet3']"
        print wb.sheetnames
        print '------carregar uma folha	    <Worksheet "Sheet3">'
        sheet = wb['Sheet3']
        print sheet
        print '#motra os dados da folha'
        print type(wb['Sheet3'])
        print "#titulo da folha			 'Sheet3'"
        print wb['Sheet3'].title
        print "#folhas ativas ( nao mostra as folhas firgens )"
        print wb.active    	    

    def test_obtendo_dados_da_folha(self):
        print ''
        print '-------------obtendo dados da folha-------------------'
        wb = openpyxl.load_workbook('G:\Meu Drive\DA  Software\Python\OpenPyXl - Execel\Execel\example.xlsx')
        sheet = wb['Sheet1']
        print sheet.title
        print "------novo titulo"
        sheet.title = 'novo titulo'	
        print sheet.title
        print "------mostra o valor da célula 	# maçã"
        print sheet['B1'].value
        print "------mostra a linha        1"
        print sheet['A1'].row
        print "------mostra a coluna       A"
        print sheet['A1'].column
        print "------cordenadas            A1"
        print sheet['A1'].coordinate 	        		

        print "------mostra a célula       <Cell Sheet1.B1>"
        print sheet.cell(row=1, column=2)
        print "------valor da celula	    ‘maçã’"
        print sheet.cell(row=1, column=2).value
        print "------retornar um bloco"
        print tuple(sheet['A1':'C3'])

    def test_escrevendo_valores_na_celula(self):
        print ''
        print '-------------escrevendo valores na celula-------------------'
        wb = openpyxl.load_workbook('G:\Meu Drive\DA  Software\Python\OpenPyXl - Execel\Execel\example.xlsx')
        sheet = wb['Sheet3']
        sheet['A1'] = 'Hello world!'
        print sheet['A1'].value


    def test_convertendo_entre_letras_e_numeros_de_colunas(self):
        print ''
        print '--------------convertendo entre letras e numeros de colunas---------------------'
        from openpyxl.utils import get_column_letter, column_index_from_string
        wb = openpyxl.load_workbook('G:\Meu Drive\DA  Software\Python\OpenPyXl - Execel\Execel\example.xlsx')
        sheet = wb['Sheet3']
        print '------ C'
        print get_column_letter(sheet.max_column)
        print '------ 27'
        print column_index_from_string('AA')


    def test_criando_folhas(self):
        print ''
        print '-------------- criando folhas ------------------'
        wb = openpyxl.Workbook()
        wb.create_sheet()				#<Worksheet "Sheet1">
        wb.create_sheet(index=1, title='segundo Sheet')	#<Worksheet "First Sheet">
        print wb.sheetnames

    def test_remover(self):
        print ''
        print '-------------- remover folhas ------------------'
        wb = openpyxl.Workbook()
        wb.create_sheet()				#<Worksheet "Sheet1">
        wb.create_sheet(index=1, title='novo Sheet')	#<Worksheet "First Sheet">
        print wb.sheetnames
        del wb['novo Sheet']
        print wb.sheetnames

    def test_salvar(self):
        print ''
        print '--------------- salvar folha -----------------'
        #wb.save('G:\\updatedProduceSales.xlsx')

    def test_obtendo_linha_e_coluna(self):
        print ''
        print '--------------obtendo linha e coluna---------------------'
        from openpyxl.utils import get_column_letter, column_index_from_string
        wb = openpyxl.load_workbook('G:\Meu Drive\DA  Software\Python\OpenPyXl - Execel\Execel\example.xlsx')
        sheet = wb['Sheet1']
        for i in range(1,8,2):
            print i, sheet.cell(row=i, column=2).value

        print '------- bloco -----'
        for linhas_da_delula in sheet['A1':'C3']:
            for celula in linhas_da_delula:
                print celula.coordinate, celula.value 
            print '--- END OF ROW ---'

    def test_atualizando_planilha(self):
        print ''
        wb = openpyxl.load_workbook('G:\Meu Drive\DA  Software\Python\OpenPyXl - Execel\Execel\produceSales.xlsx')
        sheet = wb['Sheet']

        PRICE_UPDATES = {'Garlic': 3.07,'Celery': 1.19,'Lemon': 1.27}
        for rowNum in range(2, sheet.max_row):	
            produceName = sheet.cell(row=rowNum, column=1).value
            if produceName in PRICE_UPDATES:
                sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
        wb.save('G:\Meu Drive\DA  Software\Python\OpenPyXl - Execel\Execel\updatedProduceSales.xlsx')

    def test_stilos(self):
        print ''
        print '------------------ Stilos -------------------'
        from openpyxl.styles import Font
        wb = openpyxl.Workbook()
        sheet = wb['Sheet']
        italico24Font = Font(size=24, italic=True)
        sheet['A1'] = 'Hello world!'
        sheet['A1'].font = italico24Font
        wb.save('G:\Meu Drive\DA  Software\Python\OpenPyXl - Execel\Execel\styled.xlsx')

    def test_pandas(self):
        print ''
        print '------------------ Pandas ------------------'
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font
        wb = openpyxl.Workbook()
        ws = wb.active

        df = pd.DataFrame(columns=['A', 'B', 'C'])
        for i in range(5):
            df.loc[i] = [2, 3, 4]

        for r in dataframe_to_rows(df, index=True, header=True):
            ws.append(r)

        italico15Font = Font(size=15, italic=True)

        for cell in ws['A'] + ws[1]:
            cell.style = 'Pandas'

        for cell in ws[2]:
            cell.font = italico15Font

        wb.save("G:\Meu Drive\DA  Software\Python\OpenPyXl - Execel\Execel\pandas_openpyxl.xlsx")
            
    def test_estilo(self):
        print '------------------- Estilos --------------------'
        print '-------- Estilos de Célula -----------'
        from openpyxl.styles import colors
        from openpyxl.styles import Font, Color
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active

        ws['A1'] = 'Hello '
        ws['B1'] = ' world!'

        ft = Font(color=colors.RED)
        ws['A1'].font = ft
        ws['B1'].font = Font(color=colors.RED, italic=True)

        print '-------- Cores básicas da fonte ---------'

        #from openpyxl.styles import Font
        from openpyxl.styles.colors import RED
        font1 = Font(color=RED)
        font2 = Font(color="FFBB00")

        ws['A3'] = 'Hello '
        ws['B3'] = ' world!'

        ws['A3'].font = font1
        ws['B3'].font = font2

        ws['B3'].font = Font(size=12)

        from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

        my_cell = ws['D2']
        my_cell.value = "My Cell"
        #----------------------------------------------------------------
        fill = PatternFill("solid", fgColor="666666")
        #fill = GradientFill(stop=("000000", "FFFFFF"))
        if fill:
           ws['D2'].fill = fill
        #----------------------------------------------------------------
        alignment = Alignment(horizontal="center", vertical="center")
        #----------------------------------------------------------------
        if alignment:
            ws.merge_cells('D2:H4')
            ws['D2'].alignment = alignment
        #----------------------------------------------------------------
        font = Font(b=True, color="FF0000")
        if font:
            ws['D2'].font = font
        #----------------------------------------------------------------
        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="ff0000")
        border = Border(top=double, left=thin, right=thin, bottom=double)
            
        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)

        rows = ws['D2:H4']

        for cell in rows[0]:
            cell.border = cell.border + top
            
        for cell in rows[-1]:
            cell.border = cell.border + bottom
        
        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = l.border + left
            r.border = r.border + right
        #----------------------------------------------------------------

        from openpyxl.styles import NamedStyle, Font, Border, Side
        highlight = NamedStyle(name="highlight")
        highlight.font = Font(bold=True, size=20)
        bd = Side(style='thick', color="000000")
        highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)

        wb.add_named_style(highlight)

        ws['A5'] = 'valor'
        ws['B5'] = 'valor'

        ws['A5'].style = highlight

        ws['B5'].style = 'highlight'
        
            
        wb.save("G:\Meu Drive\DA  Software\Python\OpenPyXl - Execel\Execel\estilos.xlsx")

    def test_imagem(self):
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        img = openpyxl.drawing.image.Image('G:\Meu Drive\DA  Software\Python\OpenPyXl - Execel\Execel\mesa_11_etapa_1.png')
        ws.add_image(img,'H5')
        wb.save("G:\Meu Drive\DA  Software\Python\OpenPyXl - Execel\Execel\imagem.xlsx")
        

if __name__ == "__main__":
    print('____Teste da classe OpenPyXl')
    unittest.main()
